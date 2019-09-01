from urllib.request import urlopen
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
#파일이름이 angel.xlsx인 파일을 불러옵니다.
#emergency=load_workbook(filename='긴급상황.xlsx')

#url 계속 바꿔주고
#file_name 새롭게 부여해야함

html = urlopen("https://phrase.dict.naver.com/detail.nhn?bigCategoryNo=1&middleCategoryNo=19&smallCategoryNo=141&targetLanguage=en")

bsObject = BeautifulSoup(html, "html.parser")

print(bsObject.head.title) # 웹 문서 전체가 출력됩니다.
nameList = bsObject.findAll("span", {"class":"info_txt"})
nameList_en = bsObject.findAll("span", {"class":"info_txt2"})



#파일 이름을 정하고, 데이터를 넣을 시트를 활성화합니다.
sheet1 = wb.active
file_name = 'save.xlsx'

#시트의 이름을 정합니다.
sheet1.title = '테스트용'


#cell 함수를 이용해 넣을 데이터의 행렬 위치를 지정해줍니다.
row_num=1
row_num2 =1
for name in nameList :

    sheet1.cell(row=row_num, column=1).value = name.get_text()
    row_num+=1

for name_en in nameList_en :
    print(name_en)

    sheet1.cell(row=row_num2, column=2).value = name_en.get_text()
    row_num2+=1


wb.save(filename=file_name)